#Main File for Data Visualizer and File Converter
#Author: Jake Collins
#Built using Python 3.7.6 64-bit
#Start Date:2/5/2020 9:30PM CST
#End Date for initial full working version: 2/10/2020 8:00PM CST
#Approximate amount of time spent until completion: 25 Hours
#Additional Comments will be added at a later time
#See the Python Crash Course 2nd Edition folder
import csv #For working with CSV files
import matplotlib.pyplot as plt #Used for making plots of data
from datetime import datetime #Used for grabbing and formatting dates in readable and appropriate manners
import re #Regex
import openpyxl #For working with XLSX files
from mod import Mod #Modular arithmetic
print('Currently files are suggested to be in the same folder as this Python File as files outside the folder this is put in have not been tested')
FileName = input("What file would you like to work with?")
#May need to add more checkers for different file types that come up
FileType = ''
TextFileRegex = re.compile(r'\.txt$')
CSVFileRegex = re.compile(r'\.csv$')
ExcelFileRegex = re.compile(r'\.xlsx$')
TextFileChecker = TextFileRegex.search(FileName)
CSVFileChecker = CSVFileRegex.search(FileName)
ExcelFileChecker = ExcelFileRegex.search(FileName)
if (TextFileChecker != None):
    FileType = TextFileChecker.group()
elif (CSVFileChecker != None):
    FileType = CSVFileChecker.group()
elif (ExcelFileChecker != None):
    FileType = ExcelFileChecker.group()
else:
    FileType = 'File type is not currently supported'
FileNameRaw = FileName.rsplit(FileType, 1)

#This block will be turned into the DataExtraction Class with its own file DataExtraction.py
if (FileType == '.txt'):
    FileOpener = open(FileName, 'r')
    FileContent = FileOpener.read()
    TextRegex = re.compile(r'.*\n')
    FileOpener.close()
    matches = re.findall(TextRegex, FileContent)
    matchesFixed = []
    for match in matches:
        match = match.strip('\n')
        matchesFixed.append(match)
elif (FileType == '.csv'):
    with open(FileName) as f:
        reader = csv.reader(f)
        header_row = next(reader)
        column_number = []
        data = list(reader)
        row_count = len(data)
        for index, column_header in enumerate(header_row):
            column_number.append(index)
        for i in range (0,row_count):
            for j in range(0,len(column_number)):
                print(data[i][j]) #This prints the data by row as found in the CSV file
elif (FileType == '.xlsx'):
    SheetContent = []
    wb = openpyxl.load_workbook(FileName)
    sheet = wb.active
    max_row=sheet.max_row
    max_column=sheet.max_column
    for i in range(1,max_row+1):
        for j in range(1,max_column+1):         
            cell_obj=sheet.cell(row=i,column=j)
            print(cell_obj.value) #Debugging Line
            SheetContent.append(cell_obj.value)
else:
    print ('File type is not currently supported')

print('Currently Support File Conversions: .txt to .csv, .csv to .txt, .csv to .xlsx, .xlsx to .csv (Just hit enter to skip this)')
convert_to = input("What file type would you like to convert to?")
#Currently Support File Conversions: .txt to .csv, .csv to .txt, .csv to .xlsx, .xlsx to .csv
if (convert_to == '.txt'):
    if (FileType == '.txt'):
        print('You already have a text file!')
    elif (FileType == '.csv'):
        FileOpener = open(FileNameRaw[0] +'.txt', 'w')
        for i in range (0,row_count):
            for j in range(0,len(column_number)):
                FileContent = FileOpener.write(str(data[i][j]) + ',')
            FileOpener.write('\n')
        FileOpener.close()
    else:
        print('File type is not currently supported')

elif (convert_to == '.csv'):
    if (FileType == '.txt'):
        outputFile = open(FileNameRaw[0] +'.csv','w',newline='\n')
        outputWriter = csv.writer(outputFile, delimiter = '\n')
        outputWriter.writerows([matchesFixed])
        outputFile.close()
    elif (FileType == '.csv'):
        print('You already have a CSV file!')
    elif (FileType == '.xlsx'):
        xlsx = openpyxl.load_workbook(FileName)
        sheet = xlsx.active
        data = sheet.rows
        outputFile = open(FileNameRaw[0] + '.csv', 'w+', newline='\n')
        for row in data:
            l = list(row)
            for i in range(len(row)):
                outputFile.write(str(l[i].value) + ',')
            outputFile.write('\n')
    else:
        print('File type is not currently supported')
elif (convert_to == '.xlsx'):
    if (FileType == '.csv'):
        wb = openpyxl.Workbook()
        sheet = wb.active
        with open(FileName, 'r') as f:
            reader = csv.reader(f)
            for row in reader:
                sheet.append(row)       
        wb.save(FileNameRaw[0] + '.xlsx')
    elif (FileType == '.xlsx'):
        print('You already have an Excel file!')
    else:
        print('File type is not currently supported')
else:
    print('File type is not currently supported')

def RepresentsInt(test):
    try:
        int(test)
        return True
    except ValueError:
        return False

DateIgnoreCase = re.compile('date', re.IGNORECASE)
if (FileType == '.csv'):
    with open(FileName) as f:
        reader = csv.reader(f)
        header_row = next(reader)
        column_number = []
        data = list(reader)
        row_count = len(data)
        for index, column_header in enumerate(header_row):
            column_number.append(index)
            if (DateIgnoreCase.match(column_header)):
                DateIndex = index
        dates = []
        entries = []
        data_column = []
        for i in range (0,row_count):
            for j in range(0,len(column_number)):
                if(RepresentsInt(data[i][j]) == True):
                    data_column.append(j)
                elif(j==DateIndex):
                    dates.append(datetime.strptime(data[i][j], '%Y-%m-%d'))
    data_column = list(dict.fromkeys(data_column))
    for i in range (0,row_count):
        for j in range(0,len(column_number)):
            if (j in data_column):
                entries.append(data[i][j])
    column_checks = []
    modular_factor = int(len(entries)/row_count)
    for w in range (0,len(entries)):
        column_check = Mod(w,modular_factor)
        column_checks.append(column_check)
    column_checks = list(dict.fromkeys(column_checks))
    listofcolumns = []
    for y in range (0,modular_factor):
        temporary_column = []
        for(i,item)in enumerate(entries, start=1):
            if (Mod(i,modular_factor)==column_checks[y]):
                temporary_column.append(item)
                if (len(temporary_column) == len(entries)/len(column_checks)):
                    listofcolumns.append(temporary_column)
        #Not particularly sure why but the indexes seem to start from the last column and to get the rest you must decrement or go into negative numbers 
    #for z in range (0,len(listofcolumns)): #Although from previous testing this should be decrementing it doesn't
        #print(z) #This goes up from 0 not down
    for a in range (0,len(listofcolumns)):
        for b in range (0,len(listofcolumns[a])):
            listofcolumns[a][b] = int(listofcolumns[a][b])
    plt.style.use('seaborn')
    fig, ax = plt.subplots()
    for c in range (0,len(listofcolumns)):
        ax.plot(dates, listofcolumns[c], c='red')
    plt.title("Graph of data found in file " + str(FileName), fontsize=24)
    plt.xlabel('',fontsize=16)
    fig.autofmt_xdate() 
    plt.ylabel('',fontsize=16)
    plt.tick_params(axis='both', which='major', labelsize=16)
plt.show()


